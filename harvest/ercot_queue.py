"""
ERCOT Generator Interconnection Status (GIS) queue harvester.

Harvests generator interconnection projects in the HOUSTON load zone with
a signed Interconnection Agreement (IA) from ERCOT's monthly GIS report.

Access pattern (per live-site inspection 2026-05-02):
  - Document list API (no auth required):
    https://www.ercot.com/misapp/servlets/IceDocListJsonWS?reportTypeId=15933
    Returns JSON with the latest GIS Report docId.
  - Download URL (no auth required):
    https://www.ercot.com/misdownload/servlets/mirDownload?doclookupId={docId}
    Returns the XLSX file directly (200 OK, no login required).

XLSX structure (GIS_Report_April2026.xlsx):
  Sheet "Project Details - Large Gen":
    Notes/intro at rows 7-29.
    Header row 31 (1-based): INR | Project Name | GIM Study Phase |
      Interconnecting Entity | POI Location | County | CDR Reporting Zone |
      Projected COD | Fuel | Technology | Capacity (MW) | ...
    Multi-line sub-headers at rows 32-35.
    Data rows start at row 36.
    CDR Reporting Zone for Houston metro area = "HOUSTON".
    IA-signed indicator: "IA" appears in GIM Study Phase text.

  Sheet "Project Details - Small Gen":
    Similar layout; column 3 is IA-date (datetime) not a phase string.
    Data rows found by scanning for INR-pattern cell in column 0.
    All rows in this sheet have signed IA (per sheet scope).

Filter: CDR Reporting Zone == "HOUSTON" (col index 6) for Large Gen.
        "IA" in GIM Study Phase string (col index 2) for Large Gen.
        CDR Reporting Zone == "HOUSTON" (col index 6) for Small Gen.

Record mapping:
  name           = Interconnecting Entity (col 3)
  description    = "{Project Name} — {Fuel}/{Technology}, {Capacity} MW"
  location_raw   = County (col 5)
  tags           = [fuel_type_label]
  extra          = {inr, project_name, fuel, technology, capacity_mw,
                    projected_cod (ISO string), zone}

Downloaded XLSX saved to data/raw/ercot/{friendly_name}.xlsx for
record-keeping and offline analysis.

Expected yield: 100-200 (67 Houston Large Gen IA-signed at Apr 2026;
plus Small Gen additions and future queue growth).
"""
from __future__ import annotations

import io
import logging
import os
import re
from datetime import datetime
from pathlib import Path
from typing import ClassVar

import requests

from harvest.base import BaseHarvester, RawCompanyRecord

logger = logging.getLogger(__name__)

_DOC_LIST_URL = (
    "https://www.ercot.com/misapp/servlets/IceDocListJsonWS"
    "?reportTypeId=15933&_=1"
)
_DOWNLOAD_URL = (
    "https://www.ercot.com/misdownload/servlets/mirDownload"
    "?doclookupId={doc_id}"
)
_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
}

_RAW_DIR = Path(__file__).parent.parent / "data" / "raw" / "ercot"

# CDR Reporting Zone value for Houston load zone
_HOUSTON_ZONE = "HOUSTON"

# Fuel type → human-readable label
_FUEL_LABELS: dict[str, str] = {
    "WIN": "Wind",
    "SOL": "Solar",
    "GAS": "Natural Gas",
    "NUC": "Nuclear",
    "OTH": "Other / Storage",
    "OIL": "Oil",
    "HYD": "Hydro",
}

# INR number pattern: e.g. "21INR0012", "15INR0064b"
_INR_RE = re.compile(r"^\d{2}INR\d{4}", re.IGNORECASE)


def _get_latest_doc(session: requests.Session) -> tuple[str, str] | None:
    """Fetch the GIS report document list and return (doc_id, friendly_name).

    Returns None on error.
    """
    try:
        resp = session.get(_DOC_LIST_URL, headers=_HEADERS, timeout=15)
        resp.raise_for_status()
    except requests.RequestException as exc:
        logger.error(f"[ercot:doc-list-error] {exc}")
        raise

    docs = (
        resp.json()
        .get("ListDocsByRptTypeRes", {})
        .get("DocumentList", [])
    )
    if not docs:
        logger.warning("[ercot:no-docs] Empty document list from ERCOT MIS API")
        return None

    first = docs[0].get("Document", {})
    doc_id = first.get("DocID")
    friendly = first.get("FriendlyName", f"GIS_Report_{doc_id}")
    return doc_id, friendly


def _download_xlsx(
    session: requests.Session, doc_id: str
) -> bytes:
    """Download the GIS XLSX and return raw bytes."""
    url = _DOWNLOAD_URL.format(doc_id=doc_id)
    try:
        resp = session.get(url, headers=_HEADERS, timeout=60)
        resp.raise_for_status()
    except requests.RequestException as exc:
        logger.error(f"[ercot:download-error] {exc}")
        raise
    return resp.content


def _save_xlsx(content: bytes, friendly_name: str) -> None:
    """Save XLSX to data/raw/ercot/. Non-fatal if directory is not writable."""
    try:
        _RAW_DIR.mkdir(parents=True, exist_ok=True)
        path = _RAW_DIR / f"{friendly_name}.xlsx"
        path.write_bytes(content)
        logger.info(f"[ercot:saved] {path} ({len(content)} bytes)")
    except OSError as exc:
        logger.warning(f"[ercot:save-warning] Could not save XLSX: {exc}")


# ── Sheet parsing ──────────────────────────────────────────────────────────────


def _find_data_rows(
    rows: list[tuple],
) -> tuple[int | None, int | None]:
    """Scan rows to find (header_idx, first_data_idx) for one sheet.

    header_idx: row index where column 0 == "INR" (0-based).
    first_data_idx: first row index where column 0 matches INR pattern.
    """
    header_idx: int | None = None
    for i, row in enumerate(rows):
        cell0 = row[0] if row else None
        if cell0 == "INR":
            header_idx = i
        if cell0 and isinstance(cell0, str) and _INR_RE.match(cell0):
            return header_idx, i
    return header_idx, None


def _parse_sheet(
    rows: list[tuple],
    is_large_gen: bool,
) -> list[RawCompanyRecord]:
    """Parse one worksheet's rows into RawCompanyRecord instances.

    For Large Gen: filters CDR Reporting Zone == HOUSTON AND "IA" in Phase.
    For Small Gen: filters CDR Reporting Zone == HOUSTON only (all rows
    in this sheet are IA-approved by definition).

    Column layout (0-indexed, both sheets):
      0: INR
      1: Project Name
      2: GIM Study Phase (Large Gen: string) | IA Date (Small Gen: datetime)
      3: Interconnecting Entity
      4: POI Location
      5: County
      6: CDR Reporting Zone
      7: Projected COD
      8: Fuel
      9: Technology
      10: Capacity (MW)
    """
    _, first_data_idx = _find_data_rows(rows)
    if first_data_idx is None:
        logger.warning("[ercot:no-data] No data rows found in sheet")
        return []

    records: list[RawCompanyRecord] = []

    for row in rows[first_data_idx:]:
        if not row or row[0] is None:
            continue
        if not (isinstance(row[0], str) and _INR_RE.match(row[0])):
            continue  # skip totals / blank / section-break rows

        inr = str(row[0]).strip()
        project_name = str(row[1] or "").strip()
        phase_or_date = row[2]
        entity = str(row[3] or "").strip()
        county = str(row[5] or "").strip()
        zone = str(row[6] or "").strip().upper()
        cod_raw = row[7]
        fuel_code = str(row[8] or "").strip().upper()
        technology = str(row[9] or "").strip()
        capacity_raw = row[10]

        # Zone filter
        if zone != _HOUSTON_ZONE:
            continue

        # Large Gen: require "IA" in the phase string
        if is_large_gen:
            phase_str = str(phase_or_date or "")
            if "IA" not in phase_str.upper():
                continue

        if not entity:
            logger.debug(f"[ercot:skip] {inr} has no entity name — skipped")
            continue

        fuel_label = _FUEL_LABELS.get(fuel_code, fuel_code or "Unknown")
        capacity_mw = float(capacity_raw) if isinstance(capacity_raw, (int, float)) else None
        cod_iso: str | None = None
        if isinstance(cod_raw, datetime):
            cod_iso = cod_raw.date().isoformat()

        description = f"{project_name} — {fuel_label}/{technology}"
        if capacity_mw is not None:
            description += f", {capacity_mw:.1f} MW"

        records.append(
            RawCompanyRecord(
                name=entity,
                source="ERCOT Interconnection Queue",
                source_url="https://www.ercot.com/gridinfo/resource",
                description=description or None,
                website=None,
                location_raw=county or "Houston, TX",
                tags=[fuel_label] if fuel_label else [],
                extra={
                    "inr": inr,
                    "project_name": project_name,
                    "fuel": fuel_code,
                    "technology": technology,
                    "capacity_mw": capacity_mw,
                    "projected_cod": cod_iso,
                    "zone": zone,
                },
            )
        )

    return records


def parse_gis_xlsx(content: bytes) -> list[RawCompanyRecord]:
    """Parse a GIS Report XLSX and return all Houston IA-signed records.

    Parses both "Project Details - Large Gen" and "Project Details - Small Gen"
    sheets. Returns combined list.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    all_records: list[RawCompanyRecord] = []

    for sheet_name, is_large in [
        ("Project Details - Large Gen", True),
        ("Project Details - Small Gen", False),
    ]:
        if sheet_name not in wb.sheetnames:
            logger.warning(f"[ercot:no-sheet] Sheet '{sheet_name}' not found")
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        recs = _parse_sheet(rows, is_large_gen=is_large)
        logger.info(
            f"[ercot:{sheet_name.split()[-2].lower()}] "
            f"{len(recs)} Houston IA-signed records"
        )
        all_records.extend(recs)

    return all_records


# ── Harvester ──────────────────────────────────────────────────────────────────


class ErcotQueueHarvester(BaseHarvester):
    """Harvest ERCOT interconnection queue for Houston IA-signed projects.

    Fetches the latest GIS Report XLSX via the ERCOT MIS document API,
    saves it to data/raw/ercot/, and parses both Large Gen and Small Gen
    sheets filtered to HOUSTON load zone + IA milestone.
    """

    SOURCE_NAME: ClassVar[str] = "ERCOT Interconnection Queue"
    SOURCE_URL: ClassVar[str] = "https://www.ercot.com/gridinfo/resource"
    SOURCE_TYPE: ClassVar[str] = "government_filing"
    UPDATE_CADENCE: ClassVar[str] = "monthly"
    SCRAPE_METHOD: ClassVar[str] = "xlsx_download"
    AUTH_REQUIRED: ClassVar[bool] = False
    EXPECTED_YIELD: ClassVar[str] = "50-150"

    def fetch(self) -> list[RawCompanyRecord]:
        """Fetch the latest GIS Report and return Houston IA-signed records.

        Returns one RawCompanyRecord per queue entry. Empty list on error.
        """
        session = requests.Session()

        self.rate_limiter.wait()
        result = _get_latest_doc(session)
        if not result:
            return []
        doc_id, friendly_name = result
        logger.info(f"[ercot:latest] doc_id={doc_id} name={friendly_name}")

        self.rate_limiter.wait()
        content = _download_xlsx(session, doc_id)
        logger.info(
            f"[ercot:downloaded] {len(content)} bytes for {friendly_name}"
        )
        _save_xlsx(content, friendly_name)

        records = parse_gis_xlsx(content)
        logger.info(f"[ercot:done] {len(records)} total records")
        return records
